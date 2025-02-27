from tempfile import NamedTemporaryFile

from openpyxl import Workbook

from django.http import HttpResponse

def xlsx_response(data: list):
    wb = Workbook()
    for d in data:
        wc = wb.create_sheet(title=d[0][:30])
        for row in d[1:]:
            wc.append(tuple(row))
    wb.remove(wb[wb.sheetnames[0]])  # remove default empty sheet
    with NamedTemporaryFile() as tmp:
        wb.save(tmp)
        tmp.seek(0)
        stream = tmp.read()
        return HttpResponse(stream, headers={
            "Content-Type": "application/vnd.ms-excel",
            "Content-Disposition": 'attachment; filename="reg.xlsx"',
        })
