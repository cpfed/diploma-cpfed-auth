from tempfile import NamedTemporaryFile

from openpyxl import Workbook

from django.contrib import admin
from django.http import HttpResponse, HttpResponseRedirect
from django.core.exceptions import PermissionDenied
from django.urls import reverse
from django.shortcuts import render

from .models import Contest, UserContest, ContestResult

# Register your models here.

admin.site.register(ContestResult)

@admin.register(UserContest)
class UserContestAdmin(admin.ModelAdmin):
    list_filter = ["contest__name"]

@admin.register(Contest)
class ContestAdmin(admin.ModelAdmin):
    actions = ['get_registrations', 'upload_contest_results', 'register_on_contest', 'add_bulk_reg']

    @admin.action(description="Экспортировать данные зарегистрированных пользователей")
    def get_registrations(self, request, queryset):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied()
        wb = Workbook()
        for contest in queryset:
            wc = wb.create_sheet(title=contest.name)
            wc.append(contest.required_fields)
            for uc in UserContest.objects.filter(contest=contest):
                wc.append(tuple(uc.get_full_reg.values()))
        wb.remove(wb[wb.sheetnames[0]])  # remove default empty sheet
        with NamedTemporaryFile() as tmp:
            wb.save(tmp)
            tmp.seek(0)
            stream = tmp.read()
            return HttpResponse(stream, headers={
                "Content-Type": "application/vnd.ms-excel",
                "Content-Disposition": 'attachment; filename="foo.xlsx"',
            })

    @admin.action(description="Загрузить результаты контеста")
    def upload_contest_results(self, request, queryset):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied()
        selected = queryset.values_list("pk", flat=True)
        if len(selected) != 1:
            return render(request, 'admin/result_message.html', {'message': "Должен быть выбран только один контест"})
        return HttpResponseRedirect(reverse("upload_contest_results") + '?id=' + str(selected[0]))

    @admin.action(description="Зарегистрировать пользователей на контест")
    def register_on_contest(self, request, queryset):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied()
        selected = queryset.values_list("pk", flat=True)
        if len(selected) != 1:
            return render(request, 'admin/result_message.html', {'message': "Должен быть выбран только один контест"})
        return HttpResponseRedirect(reverse("register_on_contest") + '?id=' + str(selected[0]))

    @admin.action(description="Добавить регистрации со старого контеста на новый")
    def add_bulk_reg(self, request, queryset):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied()
        selected = queryset.values_list("pk", flat=True)
        if len(selected) != 2:
            return render(request, 'admin/result_message.html', {'message': "Должены быть выбраны ровно два контеста"})
        fr, to = sorted(selected)
        contest = Contest.objects.get(id=to)
        for uc in UserContest.objects.filter(contest__id=fr):
            UserContest.objects.get_or_create(user=uc.user, contest=contest, defaults={"additional_fields": uc.additional_fields})
        return render(request, 'admin/result_message.html', {'message': "ok"})

    def get_actions(self, request):
        actions = super().get_actions(request)
        if "delete_selected" in actions:
            del actions["delete_selected"]
        return actions
